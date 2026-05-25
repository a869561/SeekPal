"""Tests del extractor Excel (.xlsx)."""
from pathlib import Path

import openpyxl
import pytest

from app.services.rag.extractors.xlsx import XlsxExtractor


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "datos.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Producto", "Cantidad", "Precio"])
    ws.append(["Manzana", 100, 0.5])
    ws.append(["Naranja", 200, 0.8])
    wb.save(str(p))
    return p


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Hoja1"
    ws1.append(["A", "B"])
    ws1.append([1, 2])
    ws2 = wb.create_sheet("Hoja2")
    ws2.append(["X", "Y"])
    ws2.append([3, 4])
    wb.save(str(p))
    return p


def test_extracts_headers_and_rows(simple_xlsx):
    doc = XlsxExtractor().extract(simple_xlsx)
    assert "Ventas" in doc.text
    assert "Producto" in doc.text
    assert "Manzana" in doc.text
    assert "Naranja" in doc.text


def test_multi_sheet(multi_sheet_xlsx):
    doc = XlsxExtractor().extract(multi_sheet_xlsx)
    assert "Hoja1" in doc.text
    assert "Hoja2" in doc.text


def test_supported_extensions():
    exts = XlsxExtractor().supported_extensions()
    assert ".xlsx" in exts
    assert ".xls" in exts


def test_empty_sheet_omitted(tmp_path):
    p = tmp_path / "vacio.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Vacía"
    ws2 = wb.create_sheet("Datos")
    ws2.append(["Solo aquí"])
    wb.save(str(p))
    doc = XlsxExtractor().extract(p)
    assert "Vacía" not in doc.text
    assert "Solo aquí" in doc.text
